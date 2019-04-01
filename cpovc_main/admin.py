"""Main module for managing set up lists."""
import csv
import time
from django.contrib import admin
from django.http import HttpResponse
from .models import SetupGeography, SetupList


def dump_to_csv(modeladmin, request, qs):
    """
    These takes in a Django queryset and spits out a CSV file.

    Generic method for any queryset
    """
    model = qs.model
    file_id = 'CPIMS_%s_%d' % (model.__name__, int(time.time()))
    file_name = 'attachment; filename=%s.csv' % (file_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = file_name
    writer = csv.writer(response, csv.excel)

    headers = []
    for field in model._meta.fields:
        headers.append(field.name)
    writer.writerow(headers)

    for obj in qs:
        row = []
        for field in headers:
            val = getattr(obj, field)
            if callable(val):
                val = val()
            if type(val) == unicode:
                val = val.encode("utf-8")
            row.append(val)
        writer.writerow(row)
    return response
dump_to_csv.short_description = u"Dump to CSV"


def export_xls(modeladmin, request, queryset):
    """Method to export as excel."""
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=list_geo.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("List Geo")
    row_num = 0
    columns = [
        (u"ID", 2000),
        (u"Name", 6000),
        (u"Parent", 8000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.area_name,
            obj.parent_area_id,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response
export_xls.short_description = u"Export XLS"


def export_xlsx(modeladmin, request, queryset):
    """Export as xlsx."""
    import openpyxl
    from openpyxl.cell import get_column_letter
    fmt = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=fmt)
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "List Geo"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Name", 70),
        (u"Parent", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        col_width = columns[col_num][1]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = col_width

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.area_name,
            obj.parent_area_id,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"


class GeoModelAdmin(admin.ModelAdmin):
    """Admin back end for Geo data management."""

    search_fields = ['area_id', 'area_name']
    list_display = ['area_id', 'area_name', 'area_type_id', 'area_code',
                    'parent_area_id']
    readonly_fields = ['area_id']
    list_filter = ['area_type_id', 'parent_area_id']
    actions = [dump_to_csv, export_xls, export_xlsx]

admin.site.register(SetupGeography, GeoModelAdmin)


class GeneralModelAdmin(admin.ModelAdmin):
    """Admin back end for Lookup lists management."""

    search_fields = ['item_id', 'item_description', 'field_name']
    list_display = ['item_id', 'item_description', 'field_name',
                    'item_category', 'item_sub_category', 'the_order',
                    'is_void']
    readonly_fields = ['is_void']
    list_filter = ['field_name']
    actions = [dump_to_csv]


admin.site.register(SetupList, GeneralModelAdmin)
