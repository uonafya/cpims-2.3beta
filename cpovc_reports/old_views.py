"""Views for all reports."""
import os
import csv
import urllib
import string
import mimetypes
import xlwt
from datetime import datetime
# from reportlab.pdfgen import canvas
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib import messages
from django.http import JsonResponse
from .forms import CaseLoad
from .functions import (
    get_case_details, case_load_header, get_data_element,
    simple_document, draw_page, get_geo_locations)

from cpovc_registry.models import RegOrgUnit
from cpovc_registry.functions import get_contacts

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from django.conf import settings

MEDIA_ROOT = settings.MEDIA_ROOT


def reports_home(request):
    """Some default page for reports home page."""
    try:
        address = 'P.O Box %s' % ('.' * 30)
        params, location = {}, '.' * 20
        form = CaseLoad(request.user)
        if request.method == 'POST':
            doc_id = request.POST.get('id')
            doc_name = request.POST.get('name')
            org_unit_id = request.POST.get('org_unit')
            file_name = '%s' % (doc_id)
            # Organisation units details
            orgs = RegOrgUnit.objects.select_related().get(
                id=org_unit_id, is_void=False)
            org_contacts = get_contacts(org_unit_id)
            params['org_unit'] = orgs.org_unit_name
            if 'contact_CPOA' in org_contacts:
                address = org_contacts['contact_CPOA'].replace('\r\n', '<br/>')
            params['address'] = address
            # Physical location
            if 'contact_CPHA' in org_contacts:
                location = org_contacts['contact_CPHA'].replace(
                    '\r\n', '<br/>')
            params['location'] = location
            # Get geo details
            geos = get_geo_locations(org_unit_id)
            sub_county = ', '.join(geos)
            params['sub_county'] = sub_county.upper()
            simple_document(document_name=doc_name, report_name=file_name,
                            params=params)
            results = {'file_name': file_name}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
        return render(request, 'reports/reports_index.html',
                      {'form': form, 'status': 200})
    except Exception, e:
        raise e


def write_xls(data, file_name):
    """Method to write xls given data."""
    try:
        book = xlwt.Workbook()
        sheet1 = book.add_sheet("Jul-15")

        for num in range(len(data)):
            row = sheet1.row(num)
            vals = data[num]
            for index, col in enumerate(vals):
                row.write(index, col)
        xls_name = '%s/%s.xls' % (MEDIA_ROOT, file_name)
        book.save(xls_name)
    except Exception, e:
        raise e


def write_xlsx(data, file_name):
    """Method to write xls given data."""
    try:
        wb = load_workbook('%s/case_load.xltm' % (MEDIA_ROOT))
        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            if sheet != 'Graph':
                ws['C3'] = file_name
            else:
                # print 'TTTTT', ws['B2'].value
                values = Reference(
                    ws, min_col=1, min_row=1, max_col=1, max_row=10)
                chart = BarChart()
                chart.add_data(values)
                ws.add_chart(chart)
        # ws0 = wb["Jul-15"]
        # ws0.title = "July-15"
        xls_name = '%s/%s.xlsx' % (MEDIA_ROOT, file_name)
        wb.save(xls_name)

        # Test sheet rename after save
        wb1 = load_workbook(xls_name)
        ws1 = wb1["Jul-15"]
        ws1.title = "Jul-17"
        ws2 = wb1["Qtr1"]
        cell_range = ws2['C9':'O139']
        # for row in ws.iter_rows('A1:C2'):
        for row in cell_range:
            for cell in row:
                value = cell.value
                if value and "Jul-15'!" in value:
                    new_value = value.replace("Jul-15'!", "Jul-17'!")
                    print new_value
                    cell.value = new_value
        xlsm_name = '%s/%s.xlsm' % (MEDIA_ROOT, file_name)
        wb1.save(xlsm_name)
    except Exception, e:
        raise e


def write_pdf(data, file_name):
    """Method to write pdf given data."""
    try:
        pdf_name = '%s/%s.pdf' % (MEDIA_ROOT, file_name)
        doc = SimpleDocTemplate(
            pdf_name, pagesize=A4, rightMargin=30,
            leftMargin=30, topMargin=30, bottomMargin=18)
        doc.pagesize = landscape(A4)
        elements = []
        # TODO: Get this line right instead of just copying it from the
        # docs
        style = TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
             ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
             ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
             ('BACKGROUND', (0, 0), (-1, 0), colors.gray)])
        # Configure style and word wrap
        s = getSampleStyleSheet()
        s = s["BodyText"]
        s.wordWrap = 'CJK'
        data2 = [[Paragraph(cell, s) for cell in row] for row in data]
        t = Table(data2)
        t.setStyle(style)
        # Send the data and build the file
        elements.append(t)
        doc.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)

    except Exception, e:
        raise e


def reports_caseload(request):
    """Case load views."""
    results, html = {}, None
    file_name = "Test"
    data, total_rows = [], 13
    cci_name = 'Charitable Children Institution Monthly Returns (CCIs)'
    si_name = 'Statutory Institution Monthly Returns'
    try:
        form = CaseLoad(request.user)
        if request.method == 'POST':
            sub_counties = request.POST.getlist('sub_county')
            results = {'res': sub_counties}
            case_categories = get_case_details(['case_category_id'])
            case_interventions = get_case_details(['intervention_id'])

            sub_county = sub_counties[0]

            # Report variables
            today = datetime.now()
            year = today.strftime('%Y')
            month = today.strftime('%b')
            report_variables = {'year': year, 'month': month,
                                'county': 'Nairobi', 'sub_county': sub_county}
            with open('%s/%s.csv' % (MEDIA_ROOT, file_name), 'wb') as csvfile:
                csvwriter = csv.writer(
                    csvfile, delimiter=',', quotechar='"',
                    quoting=csv.QUOTE_MINIMAL)
                cnt = 0
                for case_category in case_categories:
                    cnt += 1
                    case_name = case_category.item_description
                    vals = [str(cnt), case_name]
                    vals.extend(['0' for x in range(total_rows)])
                    data.append(vals)
                    # csvwriter.writerow([cnt, case_name] + ['0'] * 13)
                cnts = 0
                for case_intervention in case_interventions:
                    cnts += 1
                    interven_name = case_intervention.item_description
                    vals = [str(cnts), interven_name]
                    vals.extend(['0' for x in range(total_rows)])
                    data.append(vals)
                    # csvwriter.writerow([cnts, interven_name] + ['0'] * 13)
                csvwriter.writerows(data)

            # Now write html from csv
            row_cnt = 0
            with open('%s/%s.csv' % (MEDIA_ROOT, file_name), 'rb') as csvdata:
                table_string = ""
                csvreader = csv.reader(csvdata)
                for row in csvreader:
                    row_cnt += 1
                    table_string += "<tr>" + \
                        "<td>" + \
                        string.join(row, "</td><td>") + \
                        "</td>" + \
                        "</tr>\n"
                    if row_cnt == 38:
                        interven_title = case_load_header('Intervention')
                        table_string += interven_title
            html = '<table class="table">'
            category_title = case_load_header(header=True)
            category_title = category_title % report_variables
            html += '%s%s' % (category_title, table_string)
            # Case summary part
            case_data = get_data_element()
            zeros = '<td>0</td>' * 13
            for data_item in case_data:
                cval = case_data[data_item]
                html += '<tr><td></td><td>%s</td>%s</tr>' % (cval, zeros)
            # CCI part
            html += "<tr><td colspan='15'>%s</td></tr>" % (cci_name)
            cci_title = case_load_header('Summary of Month')
            html += cci_title
            case_data = get_data_element('cci')
            for data_item in case_data:
                cval = case_data[data_item]
                html += '<tr><td></td><td>%s</td>%s</tr>' % (cval, zeros)
            # SI part
            html += "<tr><td colspan='15'>%s</td></tr>" % (si_name)
            si_title = case_load_header('Summary of Month')
            html += si_title
            for data_item in case_data:
                cval = case_data[data_item]
                html += '<tr><td></td><td>%s</td>%s</tr>' % (cval, zeros)
            # Other information part
            html += "<tr><td colspan='15'>Additional Information</td></tr>"
            case_data = get_data_element('more_info')
            for data_item in case_data:
                cval = case_data[data_item]
                html += '<tr><td></td><td>%s</td>%s</tr>' % (cval, zeros)
            html += '</table>'

            # Lets is for pdf
            write_pdf(data, file_name)
            # This is for xls
            write_xls(data, file_name)
            # This is for xlsx
            write_xlsx(data, file_name)

        return render(request, 'reports/case_load.html',
                      {'form': form, 'results': results,
                       'report': html, 'file_name': file_name})
    except Exception, e:
        raise e


def reports_download(request, file_name):
    """Generic method for downloading files."""
    try:
        file_path = '%s/%s' % (MEDIA_ROOT, file_name)
        fp = open(file_path, 'rb')
        response = HttpResponse(fp.read())
        fp.close()
        mime_type, encoding = mimetypes.guess_type(file_name)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        response['Content-Type'] = mime_type
        response['Content-Length'] = str(os.stat(file_path).st_size)
        if encoding is not None:
            response['Content-Encoding'] = encoding

        # To inspect details for the below code, see
        # http://greenbytes.de/tech/tc2231/
        if u'WebKit' in request.META['HTTP_USER_AGENT']:
            # Safari 3.0 and Chrome 2.0 accepts UTF-8 encoded string
            # directly.
            filename_header = 'filename=%s' % file_name.encode(
                'utf-8')
        elif u'MSIE' in request.META['HTTP_USER_AGENT']:
            # IE does not support internationalized filename at all.
            # It can only recognize internationalized URL, so we do the
            # trick via routing rules.
            filename_header = ''
        else:
            # For others like Firefox, we follow RFC2231 (encoding
            # extension in HTTP headers).
            filename_header = 'filename*=UTF-8\'\'%s' % urllib.quote(
                file_name.encode('utf-8'))
        response['Content-Disposition'] = 'attachment; ' + filename_header
    except Exception, e:
        msg = 'Error getting file - %s' % (str(e))
        messages.info(request, msg)
        results, html, file_name = {}, None, None
        form = CaseLoad()
        return render(request, 'reports/case_load.html',
                      {'form': form, 'results': results,
                       'report': html, 'file_name': file_name})
    else:
        return response
